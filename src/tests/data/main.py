#!/usr/bin/ env python3
from multiprocessing import TimeoutError
from multiprocessing.pool import ThreadPool
import os
import pathlib
import time
from typing import Union

import openpyxl


def timeout(seconds):
    # Decorator Timeout
    def decorator(function):
        def wrapper(*args, **kwargs):
            pool = ThreadPool(processes=1)
            result = pool.apply_async(function, args=args, kwds=kwargs)
            try:
                return result.get(timeout=seconds)
            except TimeoutError as e:
                return e
        return wrapper
    return decorator


class ExcelDatas(object):
    """"""
    def __init__(self, excel_file_url: str) -> None:
        """"""
        self.__excel_file_url = excel_file_url
        self.__excel_values = None
        self.__excel_formulas = None
        self.__opened_correctly = True

        self.__get_excel_datas()
        
    @property
    def opened_correctly(self):
        return self.__opened_correctly
    
    @property
    def excel_data_values(self):
        return self.__excel_values
    
    @property
    def excel_data_formulas(self):
        return self.__excel_formulas

    def show_excel_data(self, data_only=True):
        """
        """
        if self.__opened_correctly:
            excel_file = self.__excel_values
            if not data_only:
                excel_file = self.__excel_formulas

            print(excel_file)
            if excel_file.sheetnames:
                for sheet_name in excel_file.sheetnames:
                    print(sheet_name + ':')

                    for row in excel_file[sheet_name].iter_rows(values_only=True):
                        print('\t', end='')

                        items = True if [x for x in row if x != None] else False
                        if items:
                            for item in row:
                                item = (str(item) + (' ' * 50))[:20]
                                print(item, end='')

                            print()
                print()
            else:
                print()
        else:
            print('File not opened correctly!')
            print('Use ')
    
    @timeout(5)
    def __get_excel_datas(self):
        self.__excel_values = openpyxl.load_workbook(
            self.__excel_file_url, data_only=True)
        
        self.__excel_formulas = openpyxl.load_workbook(
            self.__excel_file_url, data_only=False)
        
        if isinstance(self.__excel_values, TimeoutError):
            self.__opened_correctly = False
        
        if isinstance(self.__excel_formulas, TimeoutError):
            self.__opened_correctly = False


def main():
    url = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'tests/data/1 - BANCO DE DADOS - Atualizado.xlsx')
    
    excel_file = ExcelDatas(excel_file_url=url)
    if excel_file.opened_correctly:
        excel_file.show_excel_data(data_only=True)

if __name__ == '__main__':
    main()
